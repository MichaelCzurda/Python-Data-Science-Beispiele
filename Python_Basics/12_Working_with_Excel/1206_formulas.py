#!python3
#1206_formulas.py - Formulas

#Formulas, which begin with an equal sign, can configure cells to contain
#values calculated from other cells. In this section, you’ll use the openpyxl
#module to programmatically add formulas to cells, just like any normal value.
#For example:
#sheet['B9'] = '=SUM(B1:B8)'

#A formula is set just like any other text value in a cell

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save('writeFormula.xlsx')

#You can also read the formula in a cell just as you would any value.
#However, if you want to see the result of the calculation for the formula
#instead of the literal formula, you must pass True for the data_only keyword
#argument to load_workbook(). This means a Workbook object can show either
#the formulas or the result of the formulas but not both. (But you can have
#multiple Workbook objects loaded for the same spreadsheet file.) 

#Formulas
wbFormulas = openpyxl.load_workbook('writeFormula.xlsx')
sheet = wb.active
print(sheet['A3'].value)

#Data instead of Formulas
wbDataOnly = openpyxl.load_workbook('writeFormula.xlsx', data_only=True)
sheet = wb.active
print(sheet['A3'].value)

#Data only doesn't work?
#The solution is open the xlsx file manually and close it, then click save. 
#After this operation, you can try the wbDataonly programming part and get the data 500
