#!python3
#1207_adjusting_rows_columns.py - Adjusting Rows and Columns

#In Excel, adjusting the sizes of rows and columns is as easy as clicking and
#dragging the edges of a row or column header. But if you need to set a row
#or column’s size based on its cells’ contents or if you want to set sizes in a
#large number of spreadsheet files, it will be much quicker to write a Python
#program to do it.

#Rows and columns can also be hidden entirely from view. Or they can
#be “frozen” in place so that they are always visible on the screen and appear
#on every page when the spreadsheet is printed (which is handy for headers).

#####################################
#Setting Row Height and Column Width#
#####################################

import openpyxl
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensons.xlsx')

#############################
#Merging and Unmerging Cells#
#############################

wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two mered cells'
wb.save('merged.xlsx')

#To unmerge cells, call the unmerge_cells() sheet method. Enter this into
#the interactive shell.

wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active
sheet.unmerge_cells('C5:D5')
sheet.unmerge_cells('A1:D3')
sheet['C5'] = 'unmerged'
sheet['A1'] = 'unmerged'
wb.save('unmerged.xlsx')

##############
#Freeze Panes#
##############
#For spreadsheets too large to be displayed all at once, it’s helpful to “freeze”
#a few of the top rows or leftmost columns onscreen. Frozen column or
#row headers, for example, are always visible to the user even as they scroll
#through the spreadsheet.These are known as freeze panes. In OpenPyXL,
#each Worksheet object has a freeze_panes attribute that can be set to a Cell
#object or a string of a cell’s coordinates. Note that all rows above and all
#columns to the left of this cell will be frozen, but the row and column of the
#cell itself will not be frozen.

#To unfreeze all panes, set freeze_panes to None or 'A1'

#Examples
#freeze_panes setting 			Rows and columns frozen
#sheet.freeze_panes = 'A2' 		Row 1
#sheet.freeze_panes = 'B1' 		Column A
#sheet.freeze_panes = 'C1' 		Columns A and B
#sheet.freeze_panes = 'C2' 		Row 1 and columns A and B
#sheet.freeze_panes = 'A1' or
#sheet.freeze_panes = None 		No frozen panes

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2'
wb.save('freezeExamples.xlsx')

#If you set the freeze_panes attribute to 'A2', row 1 will always be viewable, no matter where the user scrolls in the spreadsheet.


