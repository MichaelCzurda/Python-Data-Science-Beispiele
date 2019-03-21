#!python3
#1201_reading_excel_docs.py - Reading Excel Documents

#Installing the openpyxl Module and then import it
#you can check out the full documentation for
#OpenPyXL at http://openpyxl.readthedocs.org/.
import openpyxl

#########################
#Reading Excel Documents#
#########################
#You can either create the spreadsheet yourself or download it from 
#http://nostarch.com/automatestuff/. -> example.xlsx

#######################################
#Opening Excel Documents with OpenPyXL#
#######################################
wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))
#<class 'openpyxl.workbook.workbook.Workbook'>

#The openpyxl.load_workbook() function takes in the filename and returns
#a value of the workbook data type. This Workbook object represents the Excel
#file, a bit like how a File object represents an opened text file

#Remember that example.xlsx needs to be in the current working directory in order 
#for you to work with it. You can find out what the current
#working directory is by importing os and using os.getcwd(), and you can
#change the current working directory using os.chdir().

##################################
#Getting sheets from the workbook#
##################################
#You can get a list of all the sheet names in the workbook by calling the wb.sheetnames
print(wb.sheetnames)
#['Sheet1', 'Sheet2', 'Sheet3']
sheet = wb['Sheet3']
print(sheet)
#<Worksheet "Sheet3">
print(type(sheet))
#<class 'openpyxl.worksheet.worksheet.Worksheet'>
print(sheet.title)
#Sheet3
anothersheet = wb.active
print(anothersheet)
#<Worksheet "Sheet1">

###############################
#Getting Cells from the Sheets#
###############################
#Once you have a Worksheet object, you can access a Cell object by its name. 
sheet = wb['Sheet1']
print(sheet['A1'])
#<Cell 'Sheet1'.A1>
print(sheet['A1'].value)
#2015-04-05 13:34:02
print(type(sheet['A1'].value))
#<class 'datetime.datetime'> -> automatically interprets value as datetime
c = sheet['B1']
print(c.value)
#Apples
print('Row '+ str(c.row) + ', Column '+str(c.column) + ' is ' + c.value)
#Row 1, Column B is Apples
print('Cell ' + str(c.coordinate) + ' is ' +c.value)
#Cell B1 is Apples

#Specifying a column by letter can be tricky to program, especially
#because after column Z, the columns start by using two letters: AA, AB,
#AC, and so on. As an alternative, you can also get a cell using the sheet’s
#cell() method and passing integers for its row and column keyword arguments. The first row or column integer is 1, not 0. 

print(sheet.cell(row=1, column=2))
#<Cell 'Sheet1'.B1>
print(sheet.cell(row=1, column=2).value)
#Apples
for i in range(1,8,2):
	print(i, sheet.cell(row=i, column=2).value)
#1 Apples
#3 Pears
#5 Apples
#7 Strawberries

# using the cell() method and its keyword arguments, you can
#write a for loop to print the values of a series of cells.

#You can determine the size of the sheet with the Worksheet object’s
#max_row and max.column.
sheet = wb['Sheet1']
print(sheet.max_row)
#7
print(sheet.max_column) #returns integer
#3

###############################################
#Converting between column Letters and Numbers#
###############################################
#To convert from letters to numbers, call the openpyxl.utils.column_index_from
#_string() function. To convert from numbers to letters, call the openpyxl.utils.get_column_letter() function.

from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1))
#A
print(get_column_letter(2))
#B
print(get_column_letter(900))
#AHP
sheet = wb['Sheet1']
print(get_column_letter(sheet.max_column))
#C
print(column_index_from_string('A'))
#1
print(column_index_from_string('AA'))
#27

##########################################
#Getting Rows and Columns from the Sheets#
##########################################
#You can slice Worksheet objects to get all the Cell objects in a row, column,
#or rectangular area of the spreadsheet. Then you can loop over all the cells
#in the slice.
sheet = wb['Sheet1']
print(tuple(sheet['A1':'C3']))
#((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell 'Sheet1'.A
#2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>), (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'
#.B3>, <Cell 'Sheet1'.C3>))

for rowOfCellsObjects in sheet['A1':'C3']:
	for cellObj in rowOfCellsObjects:
		print(cellObj.coordinate, cellObj.value)
	print('---ENF OF ROW---')
#A1 2015-04-05 13:34:02
#B1 Apples
#C1 73
#---ENF OF ROW---
#A2 2015-04-05 03:41:23
#B2 Cherries
#C2 85
#---ENF OF ROW---
#A3 2015-04-06 12:46:51
#B3 Pears
#C3 14
#---ENF OF ROW---

#Here, we specify that we want the Cell objects in the rectangular area
#from A1 to C3, and we get a Generator object containing the Cell objects in
#that area. To help us visualize this Generator object, we can use tuple() on it
#to display its Cell objects in a tuple.
#This tuple contains three tuples: one for each row, from the top of the
#desired area to the bottom. Each of these three inner tuples contains the
#Cell objects in one row of our desired area, from the leftmost cell to the right.
#So overall, our slice of the sheet contains all the Cell objects in the area
#from A1 to C3, starting from the top-left cell and ending with the bottomright cell.
#To print the values of each cell in the area, we use two for loops.

#To access the values of cells in a particular row or column, you can also
#use a Worksheet object’s rows and columns attribute. 
#Example:
sheet = wb.active
print(sheet.columns) 
#<generator object Worksheet._cells_by_col at 0x000000E1B40DDBA0>
for cellObj in sheet.columns:
	for elem in cellObj:
		print(elem.value)
#2015-04-05 13:34:
#2015-04-05 03:41:
#2015-04-06 12:46:
#2015-04-08 08:59:
#2015-04-10 02:07:
#2015-04-10 18:10:
#2015-04-10 02:40:
#Apples
#Cherries
#Pears
#Oranges
#Apples
#Bananas
#Strawberries
#73
#85
#14
#52
#152
#23
#98

##########################
#Workbooks, Sheets, Cells#
##########################
#As a quick review, here’s a rundown of all the functions, methods, and data
#types involved in reading a cell out of a spreadsheet file:
#1. Import the openpyxl module.
#2. Call the openpyxl.load_workbook() function.
#3. Get a Workbook object.
#4. Call workbook.active or workbook['Sheetname']
#5. Get a Worksheet object.
#6. Use indexing or the cell() sheet method with row and column keyword arguments.
#7. Get a Cell object.
#8. Read the Cell object’s value attribute.
