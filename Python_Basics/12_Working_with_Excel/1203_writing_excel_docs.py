#!python3
#1203_writing_excel_docs.py - Writing Excel Documents

import openpyxl, os

# create a new, blank Workbook object. 
wb = openpyxl.Workbook()
print(wb.sheetnames)
#['Sheet']

#Get sheet title
sheet = wb.active
print(sheet.title)
#Sheet

#Change sheet tilte
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.sheetnames)
#['Spam Bacon Eggs Sheet']

#Any time you modify the Workbook object or its sheets and cells, the
#spreadsheet file will not be saved until you call the save() workbook method. 
#Load examples.xslx -> most be in current working directory

print(os.getcwd())
#C:\Users\michaelc.AITC\Documents\Python Scripts\ATBS\12_Working_with_Excel

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
#Save Excel
wb.save('example_copy.xlsx')
#Whenever you edit a spreadsheet you’ve loaded from a file, you should
#always save the new, edited spreadsheet to a different filename than the original.

##############################
#Creating and Removing Sheets#
##############################
#create_sheet() -> create
#remove() -> remove

wb = openpyxl.Workbook()
wb.create_sheet()
print(wb.sheetnames)
#['Sheet', 'Sheet1']
wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames)
#['First Sheet', 'Sheet', 'Sheet1']
wb.create_sheet(index=2, title='Middle Sheet')
print(wb.sheetnames)
#['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet1']

wb.remove(wb['Middle Sheet'])
wb.remove(wb['Sheet1'])
print(wb.sheetnames)
#['First Sheet', 'Sheet']
#The remove() method takes a Worksheet object, not a string of the sheet name, as its argument

#########################
#Writing Values to Cells#
#########################
#get sheet object
sheet = wb['Sheet']
sheet['A1'] = 'Hello World'
print(sheet['A1'].value)
#Hello World
